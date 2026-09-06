#!/usr/bin/env python3
"""Curate provider acquisition records without modifying the source workbooks.

Requires openpyxl. The 2014 workbook is deliberately not used: those records
are a different acquisition, not the 23 specimens processed in this study.
"""
from __future__ import annotations
import argparse
import csv
import hashlib
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DEFAULT = ROOT / "data/cluster_wes_retrieval/2026-09-05/ovcan_human_wes_handoff_2026-09-05"
# Explicit mappings retain biologically significant R/R2 and passage labels.
NAMES = {
    "OV1369(R2)_P66": "OV1369-R2", "OV2295_P61": "OV2295",
    "OV2295(R2)_P70": "OV2295-R2", "OV3133(R)_P71": "OV3133-R",
    "OV3133(R2)_P58": "OV3133-R2", "OV3331_P54": "OV3331",
    "TOV112D_P83": "TOV112D", "TOV1369_P65": "TOV1369",
    "TOV21G_P59": "TOV21G", "TOV2295(R)_P57": "TOV2295-R",
    "TOV2835EP_P64": "TOV2835EP", "TOV2881EP_P64": "TOV2881EP",
    "TOV2929D_P57": "TOV2929D", "TOV3121D_P68": "TOV3121D",
    "TOV3121EP_P68": "TOV3121EP", "TOV3133D_P66": "TOV3133D",
    "TOV3133G_P65": "TOV3133G", "TPV81D_23-pool": "TOV81D",
    "OV2085_P64": "OV2085", "OV3291_P46": "OV3291", "OV90_P63": "OV90",
    "TOV2414_P65": "TOV2414", "TOV3392D_P69": "TOV3392D",
}
FIELDS = {
    "provider_name": "Name", "provider_library_name": "Library Name",
    "run_type": "Run Type", "library_source": "Library Source",
    "capture_library_type": "Library Type", "sequencing_type": "Type of Sequencing",
    "adapter": "Adapter", "provider_run": "Run", "capture_bed_files": "BED Files",
    "run_start_date": "Run Start Date", "read_set_id": "Read Set Id",
    "provider_number_of_reads": "Number of Reads", "provider_number_of_cycles": "Number of Cycles",
}

def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--handoff", type=Path, default=DEFAULT)
    args = ap.parse_args()
    with (ROOT / "reports/audit_2026-09-05/wes_cluster_models.csv").open() as f:
        cohort = {r["cell_line"]: r for r in csv.DictReader(f)}
    records, sources = [], []
    for year in (2017, 2018):
        source = args.handoff / "mcgill_r004741_provenance" / f"WES {year} Mes-Masson.xlsx"
        sha = hashlib.sha256(source.read_bytes()).hexdigest()
        wb = load_workbook(source, read_only=True, data_only=True)
        sheet = wb["HiSeqRead Tab"]
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows))
        for rownum, row in enumerate(rows, 2):
            if not str(row[0] or "").strip():
                continue
            name = str(row[0])
            if name not in NAMES:
                raise ValueError(f"Unmapped provider acquisition record: {name}")
            model = NAMES[name]
            rec = {"cell_line": model, "wes_sample_id": cohort[model]["cnv_sample_id"].removesuffix("_new"),
                   "recorded_wes_passage": cohort[model]["recorded_wes_passage"],
                   "mapping_status": "candidate_alias_requires_author_confirmation" if model == "TOV81D" else "name_and_passage_match",
                   "source_file": str(source.relative_to(args.handoff)), "source_sheet": sheet.title,
                   "source_row": rownum, "source_sha256": sha}
            for target, field in FIELDS.items():
                col = headers.index(field)
                value = row[col]
                if hasattr(value, "isoformat"):
                    value = value.isoformat().split("T")[0]
                rec[target] = value
                sources.append({"cell_line": model, "field": target, "value": value,
                                "source_file": rec["source_file"], "source_sheet": sheet.title,
                                "source_cell": f"{get_column_letter(col + 1)}{rownum}", "source_sha256": sha})
            # Explicit PE100 field, not the total run-cycle count (208 or 218).
            assert rec["sequencing_type"] == "Illumina HiSeq 4000 PE100"
            assert rec["run_type"] == "PAIRED_END"
            rec["read_configuration"] = "2 x 100 bp"
            rec["provider_read_count_unit"] = "provider-reported Number of Reads; not used as pipeline read denominator"
            rec["mapping_note"] = (
                "Provider sequencing record TPV81D_23-pool; Sample Tab contains TOV81D_P23 and TOV81D_P23_-2. "
                "Alias and pooling composition need author confirmation. Pipeline before-fastp paired-read count "
                "65,050,584 differs from provider count 65,085,803; proximity is not identity proof."
                if model == "TOV81D" else "")
            records.append(rec)
        wb.close()
    assert len(records) == 23 and {r["cell_line"] for r in records} == set(cohort)
    assert len({r["cell_line"] for r in records}) == 23
    out = ROOT / "output"
    out.mkdir(exist_ok=True)
    for name, rows in (("wes_acquisition_records.csv", records), ("wes_acquisition_sources.csv", sources)):
        with (out / name).open("w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0]))
            writer.writeheader()
            writer.writerows(sorted(rows, key=lambda r: r["cell_line"]))
    summary = {
        "provider_records": 23, "name_and_passage_matches": 22, "candidate_aliases": ["TOV81D"],
        "run_counts": {"4316 (2017-08-11)": 18, "4706 (2018-05-25)": 5},
        "sequencing_type": "Illumina HiSeq 4000 PE100", "capture": "Roche NimbleGen SeqCap EZ Exome v3",
        "adapter": "Illumina TruSeq LT", "source_workbooks_read_only": True,
        "excluded_source": "2014 acquisition workbook is not evidence for the current 23 WES specimens",
        "limitations": ["DNA extraction and detailed library preparation are not specified by these sequencing rows.",
                        "Confirm TOV81D alias and pool composition before submission.",
                        "2018 Library Tab duplicates sequencing table structure; it is not independent library QC.",
                        "Run-cycle totals are not read lengths; PE100 is explicit in Type of Sequencing."]}
    (out / "wes_acquisition_summary.json").write_text(json.dumps(summary, indent=2) + "\n")
    print(json.dumps(summary, indent=2))

if __name__ == "__main__":
    main()
